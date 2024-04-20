
#coding: utf-8
import os
import time
import torch
import random
import openpyxl
import argparse
import numpy as np
import torch.nn as nn
from tqdm import tqdm
import torch.optim as optim
import torch.nn.functional as F
from protonet import PrototypicalNetwork
from utils import metrics, AverageMeter, multiclass_mcc
from torch.utils.tensorboard.writer import SummaryWriter
os.environ["CUDA_VISIBLE_DEVICES"] = "0"
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
# print(torch.__version__)
# print(torch.cuda.is_available())

def format_time(seconds):
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return hours, minutes, seconds

def get_data():
    #分别生成训练，验证。测试
    pre_data = openpyxl.load_workbook('./dataset/6_mer_banlance.xlsx')#
    table = pre_data.active
    nrows = table.max_row  # 获取该sheet中的有效行数
    tezheng_dim = 4096
    feature = np.zeros((nrows, tezheng_dim))
    data_label = np.arange(nrows)

    for i in range(nrows):
        if   table.cell(i + 1, 1).value == 0:
            data_label[i] = 0
        elif table.cell(i + 1, 1).value == 1:
            data_label[i] = 1
        elif table.cell(i + 1, 1).value == 2:
            data_label[i] = 2
        elif table.cell(i + 1, 1).value == 3:
            data_label[i] = 3
        elif table.cell(i + 1, 1).value == 4:
            data_label[i] = 4

        for m in range(tezheng_dim):
            feature[i][m] = table.cell(i + 1, m + 2).value

    index0 = []
    index1 = []
    index2 = []
    index3 = []
    index4 = []
    for i in range(nrows):
        if   data_label[i] == 0:
            index0.append(i)
        elif data_label[i] == 1:
            index1.append(i)
        elif data_label[i] == 2:
            index2.append(i)
        elif data_label[i] == 3:
            index3.append(i)
        elif data_label[i] == 4:
            index4.append(i)

    test_data_indexes  = []
    val_data_indexes   = []
    train_data_indexes = []

    data_indexes = [index0, index1, index2, index3, index4]

    for i in range(len(data_indexes)): #6/10作为训练集，2/10作为验证集，2/10作为测试集
        sample_size = len(data_indexes[i]) // 10 * 2
        test_data_indexes.append(random.sample(data_indexes[i], sample_size))
        remaining_indexes = list(set(data_indexes[i]) - set(test_data_indexes[i]))
        remaining_size = len(remaining_indexes)
        val_sample_size = remaining_size // 10 * 2
        val_data_indexes.append(random.sample(remaining_indexes, val_sample_size))
        train_data_indexes.append(list(set(remaining_indexes) - set(val_data_indexes[i])))

    # for i in range(len(data_indexes)): #7/10作为训练集，1/10作为验证集，2/10作为测试集
    #     total_size = len(data_indexes[i])
    #
    #     test_sample_size = total_size // 10 * 2
    #     test_data_indexes.append(random.sample(data_indexes[i], test_sample_size))
    #
    #     remaining_indexes = list(set(data_indexes[i]) - set(test_data_indexes[i]))
    #     remaining_size = len(remaining_indexes)
    #
    #     val_sample_size = remaining_size // 10
    #     val_data_indexes.append(random.sample(remaining_indexes, val_sample_size))
    #
    #     train_data_indexes.append(list(set(remaining_indexes) - set(val_data_indexes[i])))


    train_elements = sum(len(sublist) for sublist in train_data_indexes)
    train_array = np.zeros(train_elements, dtype=int)
    index11 = 0
    for sublist in train_data_indexes:
        sublist_size = len(sublist)
        train_array[index11: index11 + sublist_size] = sublist
        index11 += sublist_size

    val_elements = sum(len(sublist) for sublist in val_data_indexes)
    val_array = np.zeros(val_elements, dtype=int)
    index22 = 0
    for sublist in val_data_indexes:
        sublist_size = len(sublist)
        val_array[index22: index22 + sublist_size] = sublist
        index22 += sublist_size

    test_elements = sum(len(sublist) for sublist in test_data_indexes)
    test_array = np.zeros(test_elements, dtype=int)
    index33 = 0
    for sublist in test_data_indexes:
        sublist_size = len(sublist)
        test_array[index33: index33 + sublist_size] = sublist
        index33 += sublist_size

    return feature,data_label,train_array,val_array,test_array

focal_loss_fn1 = nn.CrossEntropyLoss()#多分类交叉熵损失函数
class FocalLoss(torch.nn.Module):
    def __init__(self, alpha=1, gamma=2, reduction='mean'):
        super(FocalLoss, self).__init__()
        self.alpha = alpha
        self.gamma = gamma
        self.reduction = reduction

    def forward(self, logits, targets):
        ce_loss = F.cross_entropy(logits, targets, reduction='none')
        pt = torch.exp(-ce_loss)
        focal_loss = self.alpha * (1 - pt) ** self.gamma * ce_loss
        if self.reduction == 'mean':
            return torch.mean(focal_loss)
        elif self.reduction == 'sum':
            return torch.sum(focal_loss)
        else:
            return focal_loss
focal_loss_fn2 = FocalLoss(alpha=1, gamma=2, reduction='mean')
#一般可以在[0.5, 1.0, 2.0]，gamma可用的参数通常为[0.999, 0.99, 0.9]

def parse_arguments():
    argparser = argparse.ArgumentParser()
    # argparser.add_argument('--train_episodes', type=int, help='Number of batches per trainepoch cycle', default=12) #12
    # argparser.add_argument('--test_episodes', type=int, help='Number of batches per testepoch cycle', default=4) #4
    # argparser.add_argument('--val_episodes', type=int, help='Number of batches per valepoch cycle', default=6) #6
    argparser.add_argument('--input_dim', type=int, help='Model input dimensions', default=4096)  # 1360,51
    # argparser.add_argument('--learning_rate', type=float, help='learning rate', default=0.01) #0.01
    argparser.add_argument('--tensorboard', default='./run/tensorboard/',  help='') #
    argparser.add_argument('--model_savepath', default='./run/modelsave/', help='') #
    argparser.add_argument('--epochs', type=int, help='epoch number', default=50) #50
    # argparser.add_argument('--weight_decays', type=float, help='', default=1e-3) #1e-3
    argparser.add_argument('--weight_decay_step', type=int, help='', default=0) #0
    argparser.add_argument('--lr_decay_step', type=int, help='', default=1) #1
    # argparser.add_argument('--k_shot', type=int, help='K-shot', default=2) #1
    # argparser.add_argument('--patience', type=int, help='', default=9) #5

    args = argparser.parse_args()

    return args
args = parse_arguments()

def index_select(args, index, data_label, k_shot):
    label_0 = []
    label_1 = []
    label_2 = []
    label_3 = []
    label_4 = []
    for m in index:
        if   data_label[m] == 0:
            label_0.append(m)
        elif data_label[m] == 1:
            label_1.append(m)
        elif data_label[m] == 2:
            label_2.append(m)
        elif data_label[m] == 3:
            label_3.append(m)
        elif data_label[m] == 4:
            label_4.append(m)

    a = np.empty(k_shot * 5, dtype=np.int64)

    for i in range(5):
        indices = random.sample(locals()[f"label_{i}"], k_shot)
        a[i * k_shot: (i + 1) * k_shot] = indices

    return a

def _select_optimizer(args,model,learning_rate):
    optimizer = optim.Adam(model.parameters(), lr=learning_rate, weight_decay=args.weight_decay_step)
    scheduler = optim.lr_scheduler.StepLR(optimizer, step_size=args.lr_decay_step, gamma=0.1)
    return optimizer, scheduler

def eval_model(args,model, focal_loss_fn,val_index, data_feature, data_label, epoch, k_shot,val_episodes):
    model.eval()
    val_losses, accuracy, F1, Recall = [AverageMeter() for i in range(4)]
    with torch.no_grad():#没有梯度
        for batch in range(val_episodes):
            support_index = index_select(args, val_index, data_label, k_shot)
            support_x = [data_feature[iddata_feature] for iddata_feature in support_index]
            support_x = torch.tensor(support_x).float()

            leave_index = []
            for i in val_index:
                if i not in support_index:
                    leave_index.append(i)
            query_index = index_select(args, leave_index, data_label, k_shot)  # 随机选择训练节点作为测试节点
            query_x = torch.stack([torch.from_numpy(data_feature[iddata_feature]).float()
                                   for iddata_feature in query_index])  # 验证过程的查询集的数据
            query_y = torch.from_numpy(data_label[query_index]).long()  # 验证过程的查询集的标签

            output = model(support_x, query_x)
            # loss
            loss = focal_loss_fn(output, query_y).item()
            val_losses.update(loss)
            pred = torch.argmax(output, dim=-1).cpu().numpy()
            query_y_cpu = query_y.cpu()
            acc, f_s, recall = metrics(pred, query_y_cpu)
            accuracy.update(acc)
            F1.update(f_s)
            Recall.update(recall)

    # print(f'Epoch: {epoch} evaluation results: Val_loss: {val_losses.avg}, mAP: {accuracy.avg}, F1: {F1.avg}, Recall: {Recall.avg}')
    return val_losses.avg, accuracy.avg, F1.avg, Recall.avg

def train_one_epoch(args,model,focal_loss_fn,train_index,data_feature,data_label, optimizer, scheduler, epoch, total_epochs, k_shot,train_episodes):
    model.train()
    train_losses = AverageMeter()
    # progress_bar = tqdm(desc="Training [epoch X/X | episodes X/X] (loss=X.X | lr=X.X)",
    #                     bar_format="{l_bar}{r_bar}",dynamic_ncols=True)
    for batch in range(train_episodes):
        batch_src_index=index_select(args, train_index, data_label, k_shot)
        support_x = [data_feature[iddata_feature] for iddata_feature in batch_src_index]
        support_x=torch.tensor(support_x).float()
        train_test_index = []
        for i in train_index:
            if i not in batch_src_index:
                train_test_index.append(i)
        batch_src_testindex = index_select(args, train_test_index, data_label, k_shot)  # 随机选择训练节点作为测试节点
        query_x = torch.stack([torch.from_numpy(data_feature[iddata_feature]).float()
                               for iddata_feature in batch_src_testindex]) # 训练过程的查询集的数据
        query_y = torch.from_numpy(data_label[batch_src_testindex]).long() # 训练过程的查询集的标签

        output = model(support_x,query_x)
        #loss = criterion(log_p_y, batch_src_testlabel)
        loss = focal_loss_fn(output,query_y)
        # loss = F.nll_loss(log_p_y,batch_src_testlabel)#根据真实值和得到的值去调整模型的参数
        optimizer.zero_grad()#梯度置零，也就是把loss关于weight的导数变成0
        loss.backward()  # 反向传播计算参数的梯度
        optimizer.step()  # 使用优化方法进行梯度更新
        train_losses.update(loss.item())
        # progress_bar.set_description("Training [epoch %d/%d | episode %d/%d] | (loss=%2.5f | lr=%f)" %
        #         (epoch, total_epochs, batch + 1, train_episodes, loss.item(), scheduler.get_last_lr()[0]))

    scheduler.step()
    return train_losses.avg

def train(args,model,focal_loss_fn,learning_rate, k_shot, train_episodes, val_episodes,
          patience, data_feature, data_label, train_index, val_index):
    start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    # print(f'[{start_time}] Start Training...')
    optimizer, scheduler = _select_optimizer(args,model, learning_rate)
    writer = SummaryWriter(args.tensorboard)
    best_acc = 0
    early_stop_list = []
    for epoch in range(args.epochs):
        train_loss =train_one_epoch(args,model,focal_loss_fn,train_index,data_feature,data_label,
                                    optimizer, scheduler, epoch, args.epochs,k_shot,train_episodes)
        val_loss, accuracy, F1, Recall= eval_model(args,model,focal_loss_fn,val_index,data_feature,data_label,epoch,k_shot,val_episodes)
        writer.add_scalars('Loss', {'TrainLoss': train_loss, 'ValLoss': val_loss}, epoch)
        writer.add_scalars('Metrics', {'mAP': accuracy, 'F1':  F1, 'Recall': Recall}, epoch)

        # save checkpoint
        torch.save(model.state_dict(), args.model_savepath + 'last_model.pth')
        if best_acc < accuracy:
            torch.save(model.state_dict(), args.model_savepath + 'best_model.pth')
            best_acc = accuracy

        # early stop
        early_stop_list.append(val_loss)
        if len(early_stop_list) - np.argmin(early_stop_list) > patience:
            break
    writer.close()
    # end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    # print(f'[{end_time}] End of all training. The highest accuracy is {best_acc}')

def calculate_metrics_multiclass(actual_labels, predicted_labels, num_classes):
    sensitivities = np.zeros(num_classes)
    specificities = np.zeros(num_classes)

    for target_class in range(num_classes):
        true_positives = 0
        true_negatives = 0
        false_positives = 0
        false_negatives = 0

        for actual, predicted in zip(actual_labels, predicted_labels):
            if actual == target_class and predicted == target_class:
                true_positives += 1
            elif actual != target_class and predicted != target_class:
                true_negatives += 1
            elif actual != target_class and predicted == target_class:
                false_positives += 1
            elif actual == target_class and predicted != target_class:
                false_negatives += 1

        sensitivity = true_positives / (true_positives + false_negatives)
        specificity = true_negatives / (true_negatives + false_positives)

        sensitivities[target_class] = sensitivity
        specificities[target_class] = specificity

    class_metrics = {}
    for target_class in range(num_classes):
        class_metrics[target_class] = {"sensitivity": sensitivities[target_class],
                                       "specificity": specificities[target_class]}

    return class_metrics


def main():

    k_shot_list = [3]#
    learning_rate_list = [0.01]#,0.001
    episodes_list =  [[22, 11]]#
    focal_loss_ = [ focal_loss_fn1,focal_loss_fn2]#
    ED_list = ['euclidean']#, 'manhattan','cosine'
    seeds_list = [20]#, 42
    patience_list = [5]#,6,7,8,9,10

    for k_shot in k_shot_list:
        for learning_rate in learning_rate_list:
            for episodes in episodes_list:
                train_episodes,val_episodes = episodes[0], episodes[1]
                for focal_loss_fn in focal_loss_:
                    for ED in ED_list:
                        for seeds in seeds_list:
                            for patience in patience_list:

                                random.seed(seeds)
                                np.random.seed(seeds)
                                torch.manual_seed(seeds)
                                torch.cuda.manual_seed(seeds)
                                torch.cuda.manual_seed_all(seeds)

                                model = PrototypicalNetwork(args.input_dim, k_shot, ED)

                                data_feature, data_label, train_index, val_index, test_index=get_data()
                                train(args,model,focal_loss_fn,learning_rate, k_shot, train_episodes, val_episodes,
                                                patience, data_feature, data_label, train_index, val_index)

                                model.load_state_dict(torch.load('./run/modelsave/best_model.pth'))
                                model.eval()
                                val_losses, accuracy, F1, Recall = [AverageMeter() for i in range(4)]
                                actual_labels = []
                                predicted_labels = []
                                with torch.no_grad():#没有梯度

                                    test_episodes = 40 // (2 * k_shot)
                                    for batch in range(test_episodes):
                                        support_index = index_select(args, test_index, data_label, k_shot)
                                        support_x = [data_feature[iddata_feature] for iddata_feature in support_index]
                                        support_x = torch.tensor(support_x).float()
                                        leave_index = []
                                        for i in test_index:
                                            if i not in support_index:
                                                leave_index.append(i)
                                        query_index = index_select(args, leave_index, data_label, k_shot)  # 随机选择训练节点作为测试节点
                                        query_x = torch.stack([torch.from_numpy(data_feature[iddata_feature]).float()
                                                    for iddata_feature in query_index]) # 验证过程的查询集的数据
                                        query_y = data_label[query_index]# 验证过程的查询集的标签

                                        output = model(support_x, query_x)
                                        pred_test = torch.argmax(output, dim=-1).cpu().numpy()
                                        actual_labels.append(query_y.tolist())
                                        predicted_labels.append(pred_test.tolist())
                                        acc, f_s, recall = metrics(pred_test, query_y)
                                        accuracy.update(acc)
                                        F1.update(f_s)
                                        Recall.update(recall)
                                        test_index = [ x for x in test_index if x not in support_index ]
                                        test_index = [ x for x in test_index if x not in query_index ]

                                # 迭代处理每组标签并添加到列表中
                                actual_labels = [item for sublist in actual_labels for item in sublist]
                                predicted_labels = [item for sublist in predicted_labels for item in sublist]
                                # 计算每个类别的灵敏度和特异性
                                class_metrics = calculate_metrics_multiclass(actual_labels, predicted_labels, 5)
                                mcc_per_class = multiclass_mcc(actual_labels, predicted_labels)

                                if accuracy.avg >= 0:
                                    print('k_shot:{}, learning_rate:{}, train_episodes:{}, val_episodes:{}, test_episodes:{}, loss:{}, ED:{}, seeds:{}, patience:{}'
                                          ' '.format(k_shot, learning_rate, train_episodes, val_episodes, test_episodes, focal_loss_fn,ED,seeds, patience))
                                    print('test:mAP:{:.4f}, F1:{:.4f}, Recall:{:.4f}'.format(accuracy.avg, F1.avg, Recall.avg))

                                    class_metrics_str_list = []
                                    for target_class, metricss in class_metrics.items():  # 输出每个类别的结果（不换行）
                                        # mcc_value = mcc_per_class[i]
                                        class_metrics_str_list.append(f"Class{target_class}:Sn:{metricss['sensitivity']:.4f}"
                                                                      f"  Sp:{metricss['specificity']:.4f}")

                                        class_metrics_str = '   '.join(class_metrics_str_list)
                                    print(class_metrics_str)

                                    mcc_strings = [f"class{i}:{mcc:.4f}" for i, mcc in enumerate(mcc_per_class)]
                                    result_string = "   ".join(mcc_strings)
                                    print('MCC:',result_string, end=' ')
                                    print('\n')



if __name__ == '__main__':

    start_time = time.time()
    main()
    end_time = time.time()
    run_time = end_time-start_time
    end_time_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(end_time))
    print(f"结束时间: [{end_time_str}]")
    hours, minutes, seconds = format_time(run_time)
    print(f"程序运行时间：{hours}小时 {minutes}分钟 {seconds}秒")
